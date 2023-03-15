import numpy as np
import pandas as pd
import psycopg2
import random
import glob
import os
import openpyxl
import re

def grab_all_workouts(workbook):
    file=workbook
    all_sheetnames = file.sheetnames
    valid_sheetnames = [name for name in all_sheetnames if "skip" not in name]
    sheets=[file[i] for i in valid_sheetnames]
    exercises=[]
    start_row = 0

    for block in sheets:
        for row in block.iter_rows(max_col=3):
            values = []
            for cell in row:
                values.append(cell.value)
                if "workout" in str(cell.value).lower():
                    start_row = cell.row
                if start_row > 0 and cell.row >= start_row and cell.row <= start_row + 12:
                    if len(values) == 3:
                        exercises.append((values[0], values[1], values[2]))
                
    workouts = []
    current_workout = []

    start = 0
    for idx, row in enumerate(exercises):
        if row[0] is None:
            continue
        if "workout" in str(row[0]).lower():
            if current_workout:
                workouts.append(current_workout)
            current_workout = []
            start = idx
        current_workout.append(row)
    stop = idx

    if current_workout:
        workouts.append(current_workout)
        
    dfs = [pd.DataFrame(i[1:], columns=['Exercise', 'Prescription', 'Weight']) for i in workouts]
    for df in dfs:
        df['Weight'] = df['Weight'].fillna(1)
        try:
            df['Weight']=df['Weight'].astype(dtype='float64')
        except:
            pass

    
    return dfs